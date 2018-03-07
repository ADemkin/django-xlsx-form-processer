from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import column_index_from_string


def load_data_from_xlsx_file(file):
    workbook = load_workbook(file)
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    rows = sheet.rows
    
    all_data = []
    
    # skip headers
    next(rows)
    next(rows)
    
    for row in rows:
        check = row[5].value
        
        if check and 'да' in check.lower():
            cells = {
                "title":{
                    "value" :row[1].value,
                    "column":row[1].column,
                },
                "text" :{
                    "value" :row[2].value,
                    "column":row[2].column,
                },
                "row"  :row[0].row,
            }
            all_data.append(cells)
    
    return all_data


def save_data_to_xlsx(xls_data, file):
    workbook = load_workbook(file)
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    
    for line in xls_data:
        
        sheet.cell(row=line['row'],
                   column=column_index_from_string(line['title']['column'])
                   ).value = line['title']['value']
        
        sheet.cell(row=line['row'],
                   column=column_index_from_string(line['text']['column'])
                   ).value = line['text']['value']
    
    return save_virtual_workbook(workbook)


def get_form_data_from_request(request):
    lists = request.POST.lists()
    
    # filter out form data from POST dictionary
    form_data = [i[1] for i in lists if i[0] == 'form_data']
    
    NUMBER_OF_FORM_FIELDS = 5
    xlsx_edit_data = []
    
    for i in range(0, len(form_data[0]), NUMBER_OF_FORM_FIELDS):
        current_row_data = {}
        current_row_data["row"] = int(form_data[0][i])
        
        current_row_data["title"] = {
            "value" :form_data[0][i + 1],
            "column":form_data[0][i + 2],
        }
        
        current_row_data["text"] = {
            "value" :form_data[0][i + 3],
            "column":form_data[0][i + 4],
        }
        
        xlsx_edit_data.append(current_row_data)
    
    return xlsx_edit_data
